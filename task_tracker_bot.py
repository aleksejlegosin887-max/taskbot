"""
Telegram Task Tracker Bot
Бот для управления задачами команды с расширенным функционалом
Версия: 2.0 (Оптимизированная)
"""

import asyncio
import logging
import re
import sqlite3
from datetime import datetime, timedelta
from typing import Optional, List, Dict
import os

from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command, CommandObject
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, FSInputFile
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ==================== НАСТРОЙКИ ====================
BOT_TOKEN = "8360197362:AAEY_SoXEAZw0AdWijfPNIJWhZZQl--a0HE"
ADMIN_USERNAME = "chasujebezoshibochno"
DAILY_REPORT_TIME = "21:00"

# Начальный список команды
TEAM_MEMBERS = [
    "@Gde_e", 
    "@black_white_vt", 
    "@Vania3858", 
    "@Haunted_family_85", 
    "@mvp_pvz_team", 
    "@suipon192"
]

# ==================== НАСТРОЙКА ЛОГИРОВАНИЯ ====================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('bot.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ==================== СОСТОЯНИЯ FSM ====================
class TaskStates(StatesGroup):
    waiting_for_comment = State()
    waiting_for_edit = State()
    waiting_for_subtask = State()
    selecting_task = State()
    admin_selecting_task = State()
    admin_task_action = State()

# ==================== ИНИЦИАЛИЗАЦИЯ БОТА ====================
bot = Bot(token=BOT_TOKEN)
storage = MemoryStorage()
dp = Dispatcher(storage=storage)

# Глобальная переменная для хранения ID администратора
ADMIN_ID = None

# ==================== БАЗА ДАННЫХ ====================
class Database:
    def __init__(self, db_path='task_tracker.db'):
        self.db_path = db_path
        self.init_db()
    
    def get_connection(self):
        """Получение подключения к БД с правильной настройкой"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn
    
    def init_db(self):
        """Инициализация базы данных"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        # Таблица пользователей
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id_tg INTEGER UNIQUE,
                username TEXT,
                full_name TEXT,
                role TEXT CHECK(role IN ('admin', 'team'))
            )
        ''')
        
        # Таблица задач с улучшенной структурой
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                task_number TEXT UNIQUE,
                tm_id INTEGER,
                tm_username TEXT,
                task_text TEXT,
                priority TEXT CHECK(priority IN ('Высокий', 'Средний', 'Низкий')),
                deadline TEXT,
                comment TEXT,
                status TEXT CHECK(status IN ('Новая', 'В работе', 'Выполнено', 'Не выполнено', 'Просрочена')),
                created_at TEXT,
                updated_at TEXT,
                is_recurring INTEGER DEFAULT 0,
                recurring_period TEXT,
                message_id INTEGER,
                created_by INTEGER,
                FOREIGN KEY (tm_id) REFERENCES users(user_id_tg),
                FOREIGN KEY (created_by) REFERENCES users(user_id_tg)
            )
        ''')
        
        # Таблица истории задач с улучшенной структурой
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS task_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                task_id INTEGER,
                task_number TEXT,
                changed_by INTEGER,
                changed_by_username TEXT,
                action TEXT,
                comment TEXT,
                timestamp TEXT,
                FOREIGN KEY (task_id) REFERENCES tasks(id),
                FOREIGN KEY (changed_by) REFERENCES users(user_id_tg)
            )
        ''')
        
        # Таблица подзадач с улучшенной структурой
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS subtasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                task_id INTEGER,
                task_number TEXT,
                text TEXT,
                is_done INTEGER DEFAULT 0,
                FOREIGN KEY (task_id) REFERENCES tasks(id)
            )
        ''')
        
        conn.commit()
        conn.close()
        logger.info("База данных инициализирована")
    
    def add_user(self, user_id_tg: int, username: str, full_name: str, role: str):
        """Добавление или обновление пользователя"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute(
                '''INSERT INTO users (user_id_tg, username, full_name, role) 
                   VALUES (?, ?, ?, ?) 
                   ON CONFLICT(user_id_tg) DO UPDATE SET 
                   username=excluded.username, 
                   full_name=excluded.full_name,
                   role=excluded.role''',
                (user_id_tg, username, full_name, role)
            )
            conn.commit()
            logger.info(f"Пользователь добавлен/обновлен: {username} (ID: {user_id_tg}, роль: {role})")
        except Exception as e:
            logger.error(f"Ошибка при добавлении пользователя: {e}")
            conn.rollback()
        finally:
            conn.close()
    
    def get_user_by_username(self, username: str) -> Optional[Dict]:
        """Получение пользователя по username"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            # Убираем @ если он есть
            username = username.lstrip('@')
            cursor.execute('SELECT * FROM users WHERE username = ? OR username = ?', 
                         (username, f'@{username}'))
            row = cursor.fetchone()
            
            if row:
                return dict(row)
            return None
        except Exception as e:
            logger.error(f"Ошибка при получении пользователя по username: {e}")
            return None
        finally:
            conn.close()
    
    def get_user_by_tg_id(self, user_id_tg: int) -> Optional[Dict]:
        """Получение пользователя по Telegram ID"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute('SELECT * FROM users WHERE user_id_tg = ?', (user_id_tg,))
            row = cursor.fetchone()
            
            if row:
                return dict(row)
            return None
        except Exception as e:
            logger.error(f"Ошибка при получении пользователя по ID: {e}")
            return None
        finally:
            conn.close()
    
    def get_admin_id(self) -> Optional[int]:
        """Получение Telegram ID администратора"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute('SELECT user_id_tg FROM users WHERE role = ?', ('admin',))
            row = cursor.fetchone()
            
            if row:
                return row['user_id_tg']
            return None
        except Exception as e:
            logger.error(f"Ошибка при получении ID администратора: {e}")
            return None
        finally:
            conn.close()
    
    def generate_task_number(self) -> str:
        """Генерация уникального номера задачи"""
        now = datetime.now()
        date_prefix = now.strftime("TASK-%Y-%m-%d")
        
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute(
                'SELECT COUNT(*) as count FROM tasks WHERE task_number LIKE ?',
                (f"{date_prefix}%",)
            )
            count = cursor.fetchone()['count'] + 1
            return f"{date_prefix}-{count:03d}"
        except Exception as e:
            logger.error(f"Ошибка генерации номера задачи: {e}")
            # Fallback
            import random
            return f"{date_prefix}-{random.randint(100, 999)}"
        finally:
            conn.close()
    
    def create_task(self, task_data: Dict) -> str:
        """Создание новой задачи"""
        task_number = self.generate_task_number()
        now = datetime.now().isoformat()
        
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            # Получаем username исполнителя
            cursor.execute('SELECT username FROM users WHERE user_id_tg = ?', (task_data['tm_id'],))
            tm_user = cursor.fetchone()
            tm_username = tm_user['username'] if tm_user else ''
            
            # Создаем задачу
            cursor.execute('''
                INSERT INTO tasks (
                    task_number, tm_id, tm_username, task_text, priority, deadline, 
                    comment, status, created_at, updated_at, is_recurring, recurring_period, created_by
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                task_number,
                task_data['tm_id'],
                tm_username,
                task_data['task_text'],
                task_data['priority'],
                task_data.get('deadline', '-'),
                task_data.get('comment', ''),
                'Новая',
                now,
                now,
                task_data.get('is_recurring', 0),
                task_data.get('recurring_period', None),
                task_data.get('created_by', task_data['tm_id'])
            ))
            
            task_id = cursor.lastrowid
            
            # Получаем username создателя
            cursor.execute('SELECT username FROM users WHERE user_id_tg = ?', 
                         (task_data.get('created_by', task_data['tm_id']),))
            creator = cursor.fetchone()
            creator_username = creator['username'] if creator else 'Система'
            
            # Добавление в историю
            cursor.execute('''
                INSERT INTO task_history (task_id, task_number, changed_by, changed_by_username, action, timestamp)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (task_id, task_number, task_data.get('created_by', task_data['tm_id']), 
                  creator_username, 'Создана', now))
            
            conn.commit()
            logger.info(f"Задача создана: {task_number}")
            return task_number
        except Exception as e:
            logger.error(f"Ошибка создания задачи: {e}")
            conn.rollback()
            raise
        finally:
            conn.close()
    
    def get_task_by_number(self, task_number: str) -> Optional[Dict]:
        """Получение задачи по номеру"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute('SELECT * FROM tasks WHERE task_number = ?', (task_number,))
            row = cursor.fetchone()
            
            if row:
                return dict(row)
            return None
        except Exception as e:
            logger.error(f"Ошибка получения задачи {task_number}: {e}")
            return None
        finally:
            conn.close()
    
    def get_user_tasks(self, user_id_tg: int, status_filter: Optional[str] = None) -> List[Dict]:
        """Получение задач пользователя"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            if status_filter:
                cursor.execute(
                    'SELECT * FROM tasks WHERE tm_id = ? AND status = ? ORDER BY created_at DESC',
                    (user_id_tg, status_filter)
                )
            else:
                cursor.execute(
                    'SELECT * FROM tasks WHERE tm_id = ? AND status NOT IN (?, ?) ORDER BY created_at DESC',
                    (user_id_tg, 'Выполнено', 'Не выполнено')
                )
            
            rows = cursor.fetchall()
            return [dict(row) for row in rows]
        except Exception as e:
            logger.error(f"Ошибка получения задач пользователя: {e}")
            return []
        finally:
            conn.close()
    
    def get_all_tasks(self, status_filter: Optional[str] = None) -> List[Dict]:
        """Получение всех задач (для администратора)"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            if status_filter:
                cursor.execute(
                    'SELECT * FROM tasks WHERE status = ? ORDER BY created_at DESC',
                    (status_filter,)
                )
            else:
                cursor.execute('SELECT * FROM tasks ORDER BY created_at DESC')
            
            rows = cursor.fetchall()
            return [dict(row) for row in rows]
        except Exception as e:
            logger.error(f"Ошибка получения всех задач: {e}")
            return []
        finally:
            conn.close()
    
    def update_task_status(self, task_number: str, new_status: str, changed_by: int) -> bool:
        """Обновление статуса задачи"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            now = datetime.now().isoformat()
            
            # Обновляем статус
            cursor.execute(
                'UPDATE tasks SET status = ?, updated_at = ? WHERE task_number = ?',
                (new_status, now, task_number)
            )
            
            # Получаем task_id
            cursor.execute('SELECT id FROM tasks WHERE task_number = ?', (task_number,))
            task_row = cursor.fetchone()
            if not task_row:
                return False
            task_id = task_row['id']
            
            # Получаем username пользователя
            cursor.execute('SELECT username FROM users WHERE user_id_tg = ?', (changed_by,))
            user_row = cursor.fetchone()
            username = user_row['username'] if user_row else 'Неизвестный'
            
            # Добавляем в историю
            cursor.execute('''
                INSERT INTO task_history (task_id, task_number, changed_by, changed_by_username, action, timestamp)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (task_id, task_number, changed_by, username, f'Статус изменен на "{new_status}"', now))
            
            conn.commit()
            logger.info(f"Статус задачи {task_number} изменен на {new_status}")
            return True
        except Exception as e:
            logger.error(f"Ошибка обновления статуса задачи: {e}")
            conn.rollback()
            return False
        finally:
            conn.close()
    
    def add_comment(self, task_number: str, comment: str, user_id: int) -> bool:
        """Добавление комментария к задаче"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            now = datetime.now().isoformat()
            
            # Получаем task_id
            cursor.execute('SELECT id FROM tasks WHERE task_number = ?', (task_number,))
            task_row = cursor.fetchone()
            if not task_row:
                return False
            task_id = task_row['id']
            
            # Получаем username
            cursor.execute('SELECT username FROM users WHERE user_id_tg = ?', (user_id,))
            user_row = cursor.fetchone()
            username = user_row['username'] if user_row else 'Неизвестный'
            
            # Добавляем в историю
            cursor.execute('''
                INSERT INTO task_history (task_id, task_number, changed_by, changed_by_username, action, comment, timestamp)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (task_id, task_number, user_id, username, 'Комментарий', comment, now))
            
            conn.commit()
            logger.info(f"Комментарий добавлен к задаче {task_number}")
            return True
        except Exception as e:
            logger.error(f"Ошибка добавления комментария: {e}")
            conn.rollback()
            return False
        finally:
            conn.close()
    
    def update_task(self, task_number: str, task_data: Dict, changed_by: int) -> bool:
        """Обновление задачи"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            now = datetime.now().isoformat()
            
            # Обновляем задачу
            cursor.execute('''
                UPDATE tasks 
                SET task_text = ?, priority = ?, deadline = ?, comment = ?, updated_at = ?
                WHERE task_number = ?
            ''', (
                task_data['task_text'],
                task_data['priority'],
                task_data['deadline'],
                task_data.get('comment', ''),
                now,
                task_number
            ))
            
            # Получаем task_id и username
            cursor.execute('SELECT id FROM tasks WHERE task_number = ?', (task_number,))
            task_row = cursor.fetchone()
            if not task_row:
                return False
            task_id = task_row['id']
            
            cursor.execute('SELECT username FROM users WHERE user_id_tg = ?', (changed_by,))
            user_row = cursor.fetchone()
            username = user_row['username'] if user_row else 'Неизвестный'
            
            # Добавляем в историю
            cursor.execute('''
                INSERT INTO task_history (task_id, task_number, changed_by, changed_by_username, action, timestamp)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (task_id, task_number, changed_by, username, 'Задача отредактирована', now))
            
            conn.commit()
            logger.info(f"Задача {task_number} обновлена")
            return True
        except Exception as e:
            logger.error(f"Ошибка обновления задачи: {e}")
            conn.rollback()
            return False
        finally:
            conn.close()
    
    def get_task_history(self, task_number: str) -> List[Dict]:
        """Получение истории задачи"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                SELECT * FROM task_history 
                WHERE task_number = ? 
                ORDER BY timestamp ASC
            ''', (task_number,))
            
            rows = cursor.fetchall()
            return [dict(row) for row in rows]
        except Exception as e:
            logger.error(f"Ошибка получения истории задачи: {e}")
            return []
        finally:
            conn.close()
    
    def add_subtask(self, task_number: str, subtask_text: str) -> bool:
        """Добавление подзадачи"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            # Получаем task_id
            cursor.execute('SELECT id FROM tasks WHERE task_number = ?', (task_number,))
            task_row = cursor.fetchone()
            if not task_row:
                return False
            task_id = task_row['id']
            
            # Добавляем подзадачу
            cursor.execute('''
                INSERT INTO subtasks (task_id, task_number, text, is_done)
                VALUES (?, ?, ?, ?)
            ''', (task_id, task_number, subtask_text, 0))
            
            conn.commit()
            logger.info(f"Подзадача добавлена к {task_number}")
            return True
        except Exception as e:
            logger.error(f"Ошибка добавления подзадачи: {e}")
            conn.rollback()
            return False
        finally:
            conn.close()
    
    def get_subtasks(self, task_number: str) -> List[Dict]:
        """Получение подзадач"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                SELECT * FROM subtasks 
                WHERE task_number = ? 
                ORDER BY id
            ''', (task_number,))
            
            rows = cursor.fetchall()
            return [dict(row) for row in rows]
        except Exception as e:
            logger.error(f"Ошибка получения подзадач: {e}")
            return []
        finally:
            conn.close()
    
    def toggle_subtask(self, subtask_id: int) -> bool:
        """Переключение статуса подзадачи"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute('SELECT is_done FROM subtasks WHERE id = ?', (subtask_id,))
            row = cursor.fetchone()
            if not row:
                return False
            
            new_status = 0 if row['is_done'] else 1
            cursor.execute('UPDATE subtasks SET is_done = ? WHERE id = ?', (new_status, subtask_id))
            
            conn.commit()
            return True
        except Exception as e:
            logger.error(f"Ошибка переключения подзадачи: {e}")
            conn.rollback()
            return False
        finally:
            conn.close()
    
    def get_tasks_with_upcoming_deadline(self, hours: int) -> List[Dict]:
        """Получение задач с приближающимся дедлайном"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                SELECT * FROM tasks 
                WHERE status IN ('Новая', 'В работе') 
                AND deadline != '-'
                ORDER BY deadline
            ''')
            
            rows = cursor.fetchall()
            return [dict(row) for row in rows]
        except Exception as e:
            logger.error(f"Ошибка получения задач с дедлайнами: {e}")
            return []
        finally:
            conn.close()
    
    def get_statistics(self, start_date: str, end_date: str) -> Dict:
        """Получение статистики по задачам"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            # Статистика по каждому пользователю
            cursor.execute('''
                SELECT 
                    tm_username,
                    tm_id,
                    COUNT(*) as total,
                    SUM(CASE WHEN status = 'Выполнено' THEN 1 ELSE 0 END) as completed,
                    SUM(CASE WHEN status = 'Просрочена' THEN 1 ELSE 0 END) as overdue,
                    SUM(CASE WHEN status = 'Не выполнено' THEN 1 ELSE 0 END) as failed,
                    SUM(CASE WHEN status IN ('Новая', 'В работе') THEN 1 ELSE 0 END) as in_progress
                FROM tasks
                WHERE datetime(created_at) BETWEEN datetime(?) AND datetime(?)
                GROUP BY tm_username, tm_id
                ORDER BY tm_username
            ''', (start_date, end_date))
            
            rows = cursor.fetchall()
            stats = {}
            
            for row in rows:
                stats[row['tm_username']] = {
                    'tm_id': row['tm_id'],
                    'total': row['total'],
                    'completed': row['completed'],
                    'overdue': row['overdue'],
                    'failed': row['failed'],
                    'in_progress': row['in_progress']
                }
            
            return stats
        except Exception as e:
            logger.error(f"Ошибка получения статистики: {e}")
            return {}
        finally:
            conn.close()

# Инициализация базы данных
db = Database()

# ==================== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ====================

async def notify_admin(message: str):
    """Отправка уведомления администратору"""
    global ADMIN_ID
    
    try:
        if not ADMIN_ID:
            ADMIN_ID = db.get_admin_id()
        
        if ADMIN_ID:
            await bot.send_message(ADMIN_ID, message, parse_mode='HTML')
            logger.info(f"Уведомление админу отправлено (ID: {ADMIN_ID})")
        else:
            logger.warning("ID администратора не найден")
    except Exception as e:
        logger.error(f"Ошибка отправки уведомления админу: {e}")

def is_admin(user_id: int) -> bool:
    """Проверка прав администратора"""
    user = db.get_user_by_tg_id(user_id)
    return user and user['role'] == 'admin'

def format_task_message(task: Dict, include_buttons: bool = True) -> tuple:
    """Форматирование сообщения с задачей"""
    # Словари для emoji
    status_emoji = {
        'Новая': '🆕',
        'В работе': '⏳',
        'Выполнено': '✅',
        'Не выполнено': '❌',
        'Просрочена': '⚠️'
    }
    
    priority_emoji = {
        'Высокий': '🔴',
        'Средний': '🟡',
        'Низкий': '🟢'
    }
    
    # Безопасное получение значений
    status = task.get('status', 'Новая')
    priority = task.get('priority', 'Средний')
    task_text = task.get('task_text', 'Нет описания')
    deadline = task.get('deadline', '-')
    comment = task.get('comment', '')
    task_number = task.get('task_number', 'N/A')
    
    # Формируем сообщение
    message = (
        f"{status_emoji.get(status, '📌')} <b>Задача #{task_number}</b>\n"
        f"{priority_emoji.get(priority, '⚪')} Приоритет: {priority}\n\n"
        f"📋 <b>Описание:</b>\n{task_text}\n\n"
        f"⏰ <b>Дедлайн:</b> {deadline}\n"
    )
    
    if comment:
        message += f"💬 <b>Комментарий:</b> {comment}\n"
    
    message += f"\n📊 <b>Статус:</b> {status}"
    
    # Получаем подзадачи
    subtasks = db.get_subtasks(task_number)
    if subtasks:
        message += "\n\n📝 <b>Подзадачи:</b>\n"
        for subtask in subtasks:
            checkbox = "✅" if subtask['is_done'] else "☐"
            message += f"{checkbox} {subtask['text']}\n"
    
    # Создаем кнопки
    keyboard = None
    if include_buttons and status not in ['Выполнено', 'Не выполнено']:
        buttons = []
        
        if status == 'Новая':
            buttons.append([InlineKeyboardButton(text="⏳ В работе", callback_data=f"status_В работе_{task_number}")])
        elif status == 'В работе':
            # Проверяем подзадачи
            all_done = all(s['is_done'] for s in subtasks) if subtasks else True
            
            if all_done:
                buttons.append([
                    InlineKeyboardButton(text="✅ Выполнено", callback_data=f"status_Выполнено_{task_number}"),
                    InlineKeyboardButton(text="❌ Не выполнено", callback_data=f"status_Не выполнено_{task_number}")
                ])
            else:
                buttons.append([InlineKeyboardButton(text="⚠️ Завершите подзадачи", callback_data="noop")])
        
        buttons.append([InlineKeyboardButton(text="💬 Комментировать", callback_data=f"comment_{task_number}")])
        
        if subtasks:
            buttons.append([InlineKeyboardButton(text="📋 Подзадачи", callback_data=f"subtasks_{task_number}")])
        
        keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
    
    return message, keyboard

def parse_task_text(text: str) -> Optional[Dict]:
    """Парсинг текста задачи"""
    try:
        # Паттерн для разбора задачи
        pattern = r'@(\S+)\s+Приоритет:\s*(Высокий|Средний|Низкий)\s+Задача:\s*(.+?)\s+Дедлайн:\s*(.+?)\s+Комментарий:\s*(.+?)(?:\s+Повтор:\s*(ежедневно|еженедельно|ежемесячно))?$'
        
        match = re.match(pattern, text, re.IGNORECASE | re.DOTALL)
        
        if not match:
            return None
        
        username = match.group(1).strip()
        priority = match.group(2).strip()
        task_text = match.group(3).strip()
        deadline = match.group(4).strip()
        comment = match.group(5).strip()
        recurring = match.group(6).strip().lower() if match.group(6) else None
        
        # Проверяем формат дедлайна
        if deadline != '-':
            try:
                datetime.strptime(deadline, '%d.%m.%Y %H:%M')
            except ValueError:
                return None
        
        return {
            'username': f'@{username}' if not username.startswith('@') else username,
            'priority': priority,
            'task_text': task_text,
            'deadline': deadline,
            'comment': comment,
            'recurring': recurring
        }
    except Exception as e:
        logger.error(f"Ошибка парсинга задачи: {e}")
        return None

async def check_deadlines():
    """Фоновая задача проверки дедлайнов"""
    await asyncio.sleep(10)  # Задержка перед первым запуском
    
    while True:
        try:
            now = datetime.now()
            tasks = db.get_tasks_with_upcoming_deadline(2)
            
            for task in tasks:
                try:
                    deadline_str = task.get('deadline', '-')
                    if deadline_str == '-':
                        continue
                    
                    deadline = datetime.strptime(deadline_str, '%d.%m.%Y %H:%M')
                    time_diff = (deadline - now).total_seconds()
                    
                    # Напоминание за 2 часа
                    if 0 < time_diff <= 7200:  # 2 часа
                        await bot.send_message(
                            task['tm_id'],
                            f"⏰ <b>Напоминание!</b>\n\n"
                            f"Задача #{task['task_number']} должна быть выполнена через {int(time_diff/3600)} ч {int((time_diff%3600)/60)} мин\n\n"
                            f"📋 {task['task_text'][:100]}...",
                            parse_mode='HTML'
                        )
                        logger.info(f"Отправлено напоминание о задаче {task['task_number']}")
                    
                    # Просрочено
                    elif now > deadline and task['status'] not in ['Выполнено', 'Не выполнено', 'Просрочена']:
                        db.update_task_status(task['task_number'], 'Просрочена', task['tm_id'])
                        
                        await bot.send_message(
                            task['tm_id'],
                            f"⚠️ <b>Задача просрочена!</b>\n\n"
                            f"Задача #{task['task_number']}\n"
                            f"Дедлайн был: {deadline_str}\n\n"
                            f"📋 {task['task_text'][:100]}...",
                            parse_mode='HTML'
                        )
                        
                        await notify_admin(
                            f"⚠️ <b>Задача просрочена!</b>\n\n"
                            f"Задача #{task['task_number']}\n"
                            f"Исполнитель: {task.get('tm_username', 'Неизвестно')}\n"
                            f"Дедлайн был: {deadline_str}\n\n"
                            f"📋 {task['task_text'][:100]}..."
                        )
                        
                        logger.warning(f"Задача {task['task_number']} просрочена")
                
                except Exception as e:
                    logger.error(f"Ошибка проверки дедлайна задачи: {e}")
                    continue
        
        except Exception as e:
            logger.error(f"Ошибка в check_deadlines: {e}")
        
        await asyncio.sleep(600)  # Проверка каждые 10 минут

async def send_daily_report():
    """Фоновая задача отправки ежедневного отчета"""
    await asyncio.sleep(30)  # Задержка перед первым запуском
    
    report_sent_today = False
    
    while True:
        try:
            now = datetime.now()
            report_time = datetime.strptime(DAILY_REPORT_TIME, '%H:%M').time()
            
            # Сброс флага в полночь
            if now.hour == 0 and now.minute == 0:
                report_sent_today = False
            
            # Отправка отчета
            if now.time().hour == report_time.hour and now.time().minute == report_time.minute:
                if not report_sent_today:
                    # Формируем отчет за день
                    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
                    today_end = now.replace(hour=23, minute=59, second=59, microsecond=999999).isoformat()
                    
                    stats = db.get_statistics(today_start, today_end)
                    
                    if stats:
                        report = f"📊 <b>Ежедневный отчет за {now.strftime('%d.%m.%Y')}</b>\n\n"
                        
                        for username, data in stats.items():
                            report += (
                                f"👤 {username}\n"
                                f"  Всего задач: {data['total']}\n"
                                f"  ✅ Выполнено: {data['completed']}\n"
                                f"  ⚠️ Просрочено: {data['overdue']}\n"
                                f"  ❌ Не выполнено: {data['failed']}\n"
                                f"  ⏳ В работе: {data['in_progress']}\n\n"
                            )
                        
                        await notify_admin(report)
                        logger.info("Ежедневный отчет отправлен")
                    else:
                        await notify_admin(
                            f"📊 <b>Ежедневный отчет за {now.strftime('%d.%m.%Y')}</b>\n\n"
                            f"Нет данных о задачах за сегодня."
                        )
                    
                    report_sent_today = True
        
        except Exception as e:
            logger.error(f"Ошибка в send_daily_report: {e}")
        
        await asyncio.sleep(60)  # Проверка каждую минуту

def create_excel_report(stats: Dict, period: str) -> str:
    """Создание Excel отчета"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет"
        
        # Заголовок
        ws['A1'] = f"Отчет по задачам ({period})"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        # Заголовки столбцов
        headers = ['Пользователь', 'Всего задач', 'Выполнено', 'Просрочено', 'Не выполнено', 'В работе']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Данные
        row = 4
        for username, data in stats.items():
            ws.cell(row=row, column=1).value = username
            ws.cell(row=row, column=2).value = data['total']
            ws.cell(row=row, column=3).value = data['completed']
            ws.cell(row=row, column=4).value = data['overdue']
            ws.cell(row=row, column=5).value = data['failed']
            ws.cell(row=row, column=6).value = data['in_progress']
            row += 1
        
        # Автоширина столбцов
        for col in range(1, 7):
            ws.column_dimensions[chr(64 + col)].width = 20
        
        # Сохранение
        filename = f'report_{period}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(filename)
        return filename
    except Exception as e:
        logger.error(f"Ошибка создания Excel отчета: {e}")
        return None

# ==================== ОБРАБОТЧИКИ КОМАНД ====================

@dp.message(Command('start'))
async def cmd_start(message: types.Message):
    """Обработчик команды /start"""
    user_id = message.from_user.id
    username = message.from_user.username or message.from_user.first_name
    full_name = message.from_user.full_name
    
    # Проверяем, является ли пользователь администратором
    is_admin_user = username.lower() == ADMIN_USERNAME.lower()
    
    # Проверяем наличие в базе
    user = db.get_user_by_tg_id(user_id)
    
    if not user:
        # Новый пользователь
        if is_admin_user:
            role = 'admin'
            db.add_user(user_id, f'@{username}', full_name, role)
            
            global ADMIN_ID
            ADMIN_ID = user_id
            
            await message.answer(
                "✅ <b>Вы добавлены как администратор!</b>\n\n"
                "Используйте команды:\n"
                "/help - справка по командам\n"
                "/tasks - просмотр всех задач\n"
                "/report - отчеты по задачам",
                parse_mode='HTML'
            )
            logger.info(f"Администратор добавлен: @{username} (ID: {user_id})")
        elif any(f'@{username}'.lower() == member.lower() for member in TEAM_MEMBERS):
            role = 'team'
            db.add_user(user_id, f'@{username}', full_name, role)
            
            await message.answer(
                "✅ <b>Вы добавлены в команду!</b>\n\n"
                "Используйте команды:\n"
                "/help - справка по командам\n"
                "/my_tasks - ваши задачи",
                parse_mode='HTML'
            )
            logger.info(f"Член команды добавлен: @{username} (ID: {user_id})")
        else:
            await message.answer(
                "❌ У вас нет доступа к этому боту.\n\n"
                "Обратитесь к администратору для получения доступа.",
                parse_mode='HTML'
            )
            logger.warning(f"Попытка доступа от неизвестного пользователя: @{username} (ID: {user_id})")
    else:
        # Существующий пользователь
        if user['role'] == 'admin':
            await message.answer(
                "👋 <b>Добро пожаловать, администратор!</b>\n\n"
                "📝 <b>Создание задачи:</b>\n"
                "<code>@username Приоритет: Высокий Задача: Текст задачи Дедлайн: 20.02.2026 15:00 Комментарий: Комментарий</code>\n\n"
                "📊 <b>Команды:</b>\n"
                "/tasks - все задачи\n"
                "/report - отчеты\n"
                "/help - справка",
                parse_mode='HTML'
            )
        else:
            await message.answer(
                "👋 <b>Добро пожаловать!</b>\n\n"
                "📋 <b>Ваши команды:</b>\n"
                "/my_tasks - мои задачи\n"
                "/help - справка",
                parse_mode='HTML'
            )

@dp.message(Command('help'))
async def cmd_help(message: types.Message):
    """Справка по командам"""
    user = db.get_user_by_tg_id(message.from_user.id)
    
    if not user:
        await message.answer("❌ У вас нет доступа к боту")
        return
    
    if user['role'] == 'admin':
        help_text = (
            "📖 <b>Справка для администратора</b>\n\n"
            "<b>Создание задачи:</b>\n"
            "<code>@username Приоритет: Высокий/Средний/Низкий Задача: Описание задачи Дедлайн: ДД.ММ.ГГГГ ЧЧ:ММ Комментарий: Комментарий</code>\n\n"
            "<b>Повторяющиеся задачи:</b>\n"
            "Добавьте в конец: <code>Повтор: ежедневно/еженедельно/ежемесячно</code>\n\n"
            "<b>Команды:</b>\n"
            "/tasks - просмотр всех задач\n"
            "/report day/week/month - отчеты\n"
            "/history TASK-XX - история задачи\n"
            "/edit TASK-XX - редактировать задачу\n"
            "/add_team @username - добавить в команду\n"
            "/add_admin @username - добавить админа"
        )
    else:
        help_text = (
            "📖 <b>Справка</b>\n\n"
            "<b>Команды:</b>\n"
            "/my_tasks - ваши задачи\n"
            "/subtask TASK-XX текст - добавить подзадачу\n"
            "/delegate TASK-XX @user - делегировать задачу\n\n"
            "<b>Кнопки:</b>\n"
            "⏳ В работе - взять задачу\n"
            "✅ Выполнено - завершить\n"
            "❌ Не выполнено - отметить невыполненной\n"
            "💬 Комментировать - добавить комментарий"
        )
    
    await message.answer(help_text, parse_mode='HTML')

@dp.message(Command('my_tasks'))
async def cmd_my_tasks(message: types.Message, state: FSMContext):
    """Просмотр своих задач"""
    user = db.get_user_by_tg_id(message.from_user.id)
    
    if not user:
        await message.answer("❌ У вас нет доступа к боту")
        return
    
    tasks = db.get_user_tasks(message.from_user.id)
    
    if not tasks:
        await message.answer("📋 У вас пока нет активных задач")
        return
    
    response = "📋 <b>Ваши активные задачи:</b>\n\n"
    
    for i, task in enumerate(tasks, 1):
        status_emoji = {
            'Новая': '🆕',
            'В работе': '⏳',
            'Просрочена': '⚠️'
        }
        
        priority_emoji = {
            'Высокий': '🔴',
            'Средний': '🟡',
            'Низкий': '🟢'
        }
        
        response += (
            f"{i}. {status_emoji.get(task['status'], '📌')} <b>#{task['task_number']}</b>\n"
            f"   {priority_emoji.get(task['priority'], '⚪')} {task['priority']} | "
            f"⏰ {task['deadline']}\n"
            f"   📋 {task['task_text'][:50]}{'...' if len(task['task_text']) > 50 else ''}\n\n"
        )
    
    response += "\n💡 Отправьте номер задачи (1, 2, 3...) для просмотра деталей"
    
    await state.set_state(TaskStates.selecting_task)
    await state.update_data(tasks=[task['task_number'] for task in tasks])
    
    await message.answer(response, parse_mode='HTML')

@dp.message(Command('tasks'))
async def cmd_all_tasks(message: types.Message, state: FSMContext):
    """Просмотр всех задач (только для администратора)"""
    if not is_admin(message.from_user.id):
        await message.answer("❌ Эта команда доступна только администратору")
        return
    
    tasks = db.get_all_tasks()
    
    if not tasks:
        await message.answer("📋 Задач пока нет")
        return
    
    response = "📋 <b>Все задачи:</b>\n\n"
    
    for i, task in enumerate(tasks, 1):
        status_emoji = {
            'Новая': '🆕',
            'В работе': '⏳',
            'Выполнено': '✅',
            'Не выполнено': '❌',
            'Просрочена': '⚠️'
        }
        
        response += (
            f"{i}. {status_emoji.get(task['status'], '📌')} <b>#{task['task_number']}</b>\n"
            f"   👤 {task.get('tm_username', 'Неизвестно')} | "
            f"📊 {task['status']}\n"
            f"   📋 {task['task_text'][:50]}{'...' if len(task['task_text']) > 50 else ''}\n\n"
        )
    
    response += "\n💡 Отправьте номер задачи (1, 2, 3...) для управления"
    
    await state.set_state(TaskStates.admin_selecting_task)
    await state.update_data(tasks=[task['task_number'] for task in tasks])
    
    await message.answer(response, parse_mode='HTML')

@dp.message(Command('report'))
async def cmd_report(message: types.Message, command: CommandObject):
    """Генерация отчета"""
    if not is_admin(message.from_user.id):
        await message.answer("❌ Эта команда доступна только администратору")
        return
    
    args = command.args
    
    if not args or args not in ['day', 'week', 'month']:
        await message.answer(
            "📊 <b>Формат команды:</b>\n"
            "/report day - отчет за день\n"
            "/report week - отчет за неделю\n"
            "/report month - отчет за месяц",
            parse_mode='HTML'
        )
        return
    
    now = datetime.now()
    
    if args == 'day':
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = now.replace(hour=23, minute=59, second=59, microsecond=999999)
        period = f"день ({now.strftime('%d.%m.%Y')})"
    elif args == 'week':
        start_date = now - timedelta(days=now.weekday())
        start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = start_date + timedelta(days=6, hours=23, minutes=59, seconds=59)
        period = f"неделю ({start_date.strftime('%d.%m')} - {end_date.strftime('%d.%m.%Y')})"
    else:  # month
        start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        next_month = start_date.replace(day=28) + timedelta(days=4)
        end_date = next_month.replace(day=1) - timedelta(seconds=1)
        period = f"месяц ({start_date.strftime('%B %Y')})"
    
    stats = db.get_statistics(start_date.isoformat(), end_date.isoformat())
    
    if not stats:
        await message.answer(f"📊 Нет данных за {period}")
        return
    
    # Текстовый отчет
    report = f"📊 <b>Отчет за {period}</b>\n\n"
    
    for username, data in stats.items():
        report += (
            f"👤 <b>{username}</b>\n"
            f"  Всего задач: {data['total']}\n"
            f"  ✅ Выполнено: {data['completed']}\n"
            f"  ⚠️ Просрочено: {data['overdue']}\n"
            f"  ❌ Не выполнено: {data['failed']}\n"
            f"  ⏳ В работе: {data['in_progress']}\n\n"
        )
    
    await message.answer(report, parse_mode='HTML')
    
    # Excel отчет
    filename = create_excel_report(stats, period)
    if filename:
        try:
            file = FSInputFile(filename)
            await message.answer_document(file, caption=f"📊 Детальный отчет за {period}")
            os.remove(filename)
        except Exception as e:
            logger.error(f"Ошибка отправки Excel файла: {e}")

@dp.message(Command('history'))
async def cmd_history(message: types.Message, command: CommandObject):
    """Просмотр истории задачи"""
    if not command.args:
        await message.answer("❌ Укажите номер задачи: /history TASK-2026-02-16-001")
        return
    
    task_number = command.args.strip()
    task = db.get_task_by_number(task_number)
    
    if not task:
        await message.answer(f"❌ Задача {task_number} не найдена")
        return
    
    # Проверяем права доступа
    user = db.get_user_by_tg_id(message.from_user.id)
    if not user:
        await message.answer("❌ У вас нет доступа")
        return
    
    if user['role'] != 'admin' and task['tm_id'] != message.from_user.id:
        await message.answer("❌ Вы можете просматривать историю только своих задач")
        return
    
    history = db.get_task_history(task_number)
    
    if not history:
        await message.answer(f"📜 История задачи {task_number} пуста")
        return
    
    response = f"📜 <b>История задачи #{task_number}</b>\n\n"
    
    for entry in history:
        timestamp = entry['timestamp'][:16] if len(entry['timestamp']) > 16 else entry['timestamp']
        username = entry.get('changed_by_username', 'Система')
        
        response += f"🕒 {timestamp} | {username}\n"
        response += f"▪️ {entry['action']}\n"
        
        if entry.get('comment'):
            response += f"💬 {entry['comment']}\n"
        
        response += "\n"
    
    await message.answer(response, parse_mode='HTML')

@dp.message(Command('edit'))
async def cmd_edit(message: types.Message, command: CommandObject, state: FSMContext):
    """Редактирование задачи"""
    if not is_admin(message.from_user.id):
        await message.answer("❌ Эта команда доступна только администратору")
        return
    
    if not command.args:
        await message.answer("❌ Укажите номер задачи: /edit TASK-2026-02-16-001")
        return
    
    task_number = command.args.strip()
    task = db.get_task_by_number(task_number)
    
    if not task:
        await message.answer(f"❌ Задача {task_number} не найдена")
        return
    
    await state.update_data(task_number=task_number)
    await state.set_state(TaskStates.waiting_for_edit)
    
    await message.answer(
        f"📝 <b>Редактирование задачи #{task_number}</b>\n\n"
        f"Текущие данные:\n"
        f"Приоритет: {task['priority']}\n"
        f"Задача: {task['task_text']}\n"
        f"Дедлайн: {task['deadline']}\n"
        f"Комментарий: {task.get('comment', '-')}\n\n"
        f"Отправьте новое описание в формате:\n\n"
        f"Приоритет: Высокий/Средний/Низкий\n"
        f"Задача: Текст задачи\n"
        f"Дедлайн: ДД.ММ.ГГГГ ЧЧ:ММ или -\n"
        f"Комментарий: Текст\n\n"
        f"Или /cancel для отмены",
        parse_mode='HTML'
    )

@dp.message(Command('subtask'))
async def cmd_subtask(message: types.Message, command: CommandObject):
    """Добавление подзадачи"""
    if not command.args:
        await message.answer("❌ Формат: /subtask TASK-2026-02-16-001 Текст подзадачи")
        return
    
    parts = command.args.split(None, 1)
    if len(parts) < 2:
        await message.answer("❌ Формат: /subtask TASK-2026-02-16-001 Текст подзадачи")
        return
    
    task_number, subtask_text = parts
    
    task = db.get_task_by_number(task_number)
    if not task:
        await message.answer(f"❌ Задача {task_number} не найдена")
        return
    
    # Проверяем права доступа
    user = db.get_user_by_tg_id(message.from_user.id)
    if not user:
        await message.answer("❌ У вас нет доступа")
        return
    
    if user['role'] != 'admin' and task['tm_id'] != message.from_user.id:
        await message.answer("❌ Вы можете добавлять подзадачи только к своим задачам")
        return
    
    if db.add_subtask(task_number, subtask_text):
        await message.answer(f"✅ Подзадача добавлена к #{task_number}")
        
        # Уведомляем другую сторону
        if user['role'] == 'admin':
            try:
                await bot.send_message(
                    task['tm_id'],
                    f"📝 К задаче #{task_number} добавлена подзадача:\n"
                    f"☐ {subtask_text}",
                    parse_mode='HTML'
                )
            except:
                pass
        else:
            await notify_admin(
                f"📝 К задаче #{task_number} добавлена подзадача:\n"
                f"☐ {subtask_text}\n"
                f"Добавил: {user['username']}"
            )
    else:
        await message.answer("❌ Ошибка добавления подзадачи")

@dp.message(Command('delegate'))
async def cmd_delegate(message: types.Message, command: CommandObject):
    """Запрос на делегирование задачи"""
    if not command.args:
        await message.answer("❌ Формат: /delegate TASK-2026-02-16-001 @username")
        return
    
    parts = command.args.split(None, 1)
    if len(parts) < 2:
        await message.answer("❌ Формат: /delegate TASK-2026-02-16-001 @username")
        return
    
    task_number, new_user = parts
    
    task = db.get_task_by_number(task_number)
    if not task:
        await message.answer(f"❌ Задача {task_number} не найдена")
        return
    
    # Проверяем что это задача пользователя
    if task['tm_id'] != message.from_user.id:
        await message.answer("❌ Вы можете делегировать только свои задачи")
        return
    
    # Отправляем запрос администратору
    await notify_admin(
        f"🔄 <b>Запрос на делегирование</b>\n\n"
        f"Задача: #{task_number}\n"
        f"От: {task.get('tm_username', 'Неизвестно')}\n"
        f"Кому: {new_user}\n\n"
        f"📋 {task['task_text'][:100]}...\n\n"
        f"Для делегирования создайте новую задачу для {new_user}"
    )
    
    await message.answer(
        "✅ Запрос на делегирование отправлен администратору.\n"
        "Ожидайте решения."
    )

@dp.message(Command('add_team'))
async def cmd_add_team(message: types.Message, command: CommandObject):
    """Добавление пользователя в команду"""
    if not is_admin(message.from_user.id):
        await message.answer("❌ Эта команда доступна только администратору")
        return
    
    if not command.args:
        await message.answer("❌ Укажите username: /add_team @username")
        return
    
    username = command.args.strip()
    if not username.startswith('@'):
        username = f'@{username}'
    
    # Проверяем существование пользователя
    user = db.get_user_by_username(username)
    if user:
        await message.answer(f"⚠️ Пользователь {username} уже в базе (роль: {user['role']})")
        return
    
    # Добавляем в список команды
    global TEAM_MEMBERS
    if username not in TEAM_MEMBERS:
        TEAM_MEMBERS.append(username)
    
    await message.answer(
        f"✅ Пользователь {username} добавлен в список команды.\n"
        f"При первом обращении к боту будет автоматически зарегистрирован."
    )

@dp.message(Command('add_admin'))
async def cmd_add_admin(message: types.Message, command: CommandObject):
    """Добавление администратора"""
    if not is_admin(message.from_user.id):
        await message.answer("❌ Эта команда доступна только администратору")
        return
    
    if not command.args:
        await message.answer("❌ Укажите username: /add_admin @username")
        return
    
    username = command.args.strip()
    if not username.startswith('@'):
        username = f'@{username}'
    
    # Проверяем существование пользователя
    user = db.get_user_by_username(username)
    if user and user['role'] == 'admin':
        await message.answer(f"⚠️ Пользователь {username} уже является администратором")
        return
    
    await message.answer(
        f"ℹ️ Пользователь {username} будет добавлен как администратор при первом обращении к боту.\n"
        f"Попросите пользователя отправить боту /start"
    )

@dp.message(Command('cancel'))
async def cmd_cancel(message: types.Message, state: FSMContext):
    """Отмена текущей операции"""
    await state.clear()
    await message.answer("❌ Операция отменена")

# ==================== ОБРАБОТЧИКИ СООБЩЕНИЙ ====================

@dp.message(TaskStates.waiting_for_comment)
async def process_comment(message: types.Message, state: FSMContext):
    """Обработка комментария"""
    data = await state.get_data()
    task_number = data.get('task_number')
    is_admin_comment = data.get('admin_comment', False)
    
    if not task_number:
        await message.answer("❌ Ошибка: задача не найдена")
        await state.clear()
        return
    
    task = db.get_task_by_number(task_number)
    if not task:
        await message.answer(f"❌ Задача {task_number} не найдена")
        await state.clear()
        return
    
    comment_text = message.text
    
    if db.add_comment(task_number, comment_text, message.from_user.id):
        await message.answer(f"✅ Комментарий добавлен к задаче #{task_number}")
        
        # Уведомляем другую сторону
        if is_admin_comment:
            # Админ комментирует - уведомляем исполнителя
            try:
                await bot.send_message(
                    task['tm_id'],
                    f"💬 <b>Новый комментарий к задаче #{task_number}</b>\n\n"
                    f"От: администратор\n"
                    f"Комментарий: {comment_text}",
                    parse_mode='HTML'
                )
            except Exception as e:
                logger.error(f"Ошибка отправки уведомления исполнителю: {e}")
        else:
            # Исполнитель комментирует - уведомляем админа
            user = db.get_user_by_tg_id(message.from_user.id)
            username = user['username'] if user else 'Неизвестный'
            
            await notify_admin(
                f"💬 <b>Новый комментарий к задаче #{task_number}</b>\n\n"
                f"От: {username}\n"
                f"Комментарий: {comment_text}\n\n"
                f"Задача: {task['task_text'][:100]}..."
            )
    else:
        await message.answer("❌ Ошибка добавления комментария")
    
    await state.clear()

@dp.message(TaskStates.waiting_for_edit)
async def process_edit(message: types.Message, state: FSMContext):
    """Обработка редактирования задачи"""
    data = await state.get_data()
    task_number = data.get('task_number')
    
    if not task_number:
        await message.answer("❌ Ошибка: задача не найдена")
        await state.clear()
        return
    
    # Парсим данные
    pattern = r'Приоритет:\s*(Высокий|Средний|Низкий)\s+Задача:\s*(.+?)\s+Дедлайн:\s*(.+?)\s+Комментарий:\s*(.+?)$'
    match = re.match(pattern, message.text, re.IGNORECASE | re.DOTALL)
    
    if not match:
        await message.answer(
            "❌ Неверный формат. Используйте:\n\n"
            "Приоритет: Высокий/Средний/Низкий\n"
            "Задача: Текст задачи\n"
            "Дедлайн: ДД.ММ.ГГГГ ЧЧ:ММ или -\n"
            "Комментарий: Текст"
        )
        return
    
    priority = match.group(1).strip()
    task_text = match.group(2).strip()
    deadline = match.group(3).strip()
    comment = match.group(4).strip()
    
    # Проверяем дедлайн
    if deadline != '-':
        try:
            datetime.strptime(deadline, '%d.%m.%Y %H:%M')
        except ValueError:
            await message.answer("❌ Неверный формат дедлайна. Используйте: ДД.ММ.ГГГГ ЧЧ:ММ или -")
            return
    
    task_data = {
        'priority': priority,
        'task_text': task_text,
        'deadline': deadline,
        'comment': comment
    }
    
    task = db.get_task_by_number(task_number)
    if not task:
        await message.answer(f"❌ Задача {task_number} не найдена")
        await state.clear()
        return
    
    if db.update_task(task_number, task_data, message.from_user.id):
        await message.answer(f"✅ Задача #{task_number} обновлена")
        
        # Уведомляем исполнителя
        try:
            await bot.send_message(
                task['tm_id'],
                f"📝 <b>Задача #{task_number} обновлена</b>\n\n"
                f"Новые данные:\n"
                f"Приоритет: {priority}\n"
                f"Задача: {task_text}\n"
                f"Дедлайн: {deadline}\n"
                f"Комментарий: {comment}",
                parse_mode='HTML'
            )
        except Exception as e:
            logger.error(f"Ошибка уведомления при редактировании: {e}")
    else:
        await message.answer("❌ Ошибка обновления задачи")
    
    await state.clear()

@dp.message(TaskStates.selecting_task)
async def process_task_selection(message: types.Message, state: FSMContext):
    """Обработка выбора задачи пользователем"""
    try:
        task_index = int(message.text) - 1
        data = await state.get_data()
        tasks = data.get('tasks', [])
        
        if 0 <= task_index < len(tasks):
            task_number = tasks[task_index]
            task = db.get_task_by_number(task_number)
            
            if task:
                msg_text, keyboard = format_task_message(task)
                await message.answer(msg_text, reply_markup=keyboard, parse_mode='HTML')
                await state.clear()
            else:
                await message.answer("❌ Задача не найдена")
        else:
            await message.answer("❌ Неверный номер задачи")
    except ValueError:
        await message.answer("❌ Отправьте номер задачи (1, 2, 3...)")

@dp.message(TaskStates.admin_selecting_task)
async def process_admin_task_selection(message: types.Message, state: FSMContext):
    """Обработка выбора задачи администратором"""
    try:
        task_index = int(message.text) - 1
        data = await state.get_data()
        tasks = data.get('tasks', [])
        
        if 0 <= task_index < len(tasks):
            task_number = tasks[task_index]
            task = db.get_task_by_number(task_number)
            
            if task:
                msg_text, _ = format_task_message(task, include_buttons=False)
                
                # Кнопки управления для админа
                keyboard = InlineKeyboardMarkup(inline_keyboard=[
                    [
                        InlineKeyboardButton(text="✏️ Редактировать", callback_data=f"admin_edit_{task_number}"),
                        InlineKeyboardButton(text="🗑 Удалить", callback_data=f"admin_delete_{task_number}")
                    ],
                    [
                        InlineKeyboardButton(text="💬 Комментарий", callback_data=f"admin_comment_{task_number}"),
                        InlineKeyboardButton(text="📜 История", callback_data=f"admin_history_{task_number}")
                    ],
                    [
                        InlineKeyboardButton(text="🔄 Изменить статус", callback_data=f"admin_status_{task_number}"),
                        InlineKeyboardButton(text="🔙 Назад", callback_data="admin_cancel")
                    ]
                ])
                
                await message.answer(msg_text, reply_markup=keyboard, parse_mode='HTML')
                await state.update_data(selected_task=task_number)
            else:
                await message.answer("❌ Задача не найдена")
        else:
            await message.answer("❌ Неверный номер задачи")
    except ValueError:
        await message.answer("❌ Отправьте номер задачи (1, 2, 3...)")

@dp.message(F.text)
async def process_task_creation(message: types.Message):
    """Обработка создания задачи"""
    # Проверяем права
    if not is_admin(message.from_user.id):
        return
    
    # Парсим задачу
    task_data = parse_task_text(message.text)
    
    if not task_data:
        return
    
    # Проверяем существование исполнителя
    tm_user = db.get_user_by_username(task_data['username'])
    
    if not tm_user:
        await message.answer(
            f"❌ Пользователь {task_data['username']} не найден в базе.\n"
            f"Пользователь должен сначала отправить боту /start"
        )
        return
    
    # Создаем задачу
    task_dict = {
        'tm_id': tm_user['user_id_tg'],
        'task_text': task_data['task_text'],
        'priority': task_data['priority'],
        'deadline': task_data['deadline'],
        'comment': task_data['comment'],
        'is_recurring': 1 if task_data.get('recurring') else 0,
        'recurring_period': task_data.get('recurring'),
        'created_by': message.from_user.id
    }
    
    try:
        task_number = db.create_task(task_dict)
        
        await message.answer(
            f"✅ Задача создана!\n"
            f"Номер: #{task_number}\n"
            f"Исполнитель: {task_data['username']}"
        )
        
        # Отправляем задачу исполнителю
        task = db.get_task_by_number(task_number)
        msg_text, keyboard = format_task_message(task)
        
        try:
            sent_message = await bot.send_message(
                tm_user['user_id_tg'],
                msg_text,
                reply_markup=keyboard,
                parse_mode='HTML'
            )
            
            # Сохраняем message_id для возможного обновления
            conn = db.get_connection()
            cursor = conn.cursor()
            cursor.execute(
                'UPDATE tasks SET message_id = ? WHERE task_number = ?',
                (sent_message.message_id, task_number)
            )
            conn.commit()
            conn.close()
            
            logger.info(f"Задача {task_number} отправлена исполнителю")
        except Exception as e:
            logger.error(f"Ошибка отправки задачи исполнителю: {e}")
            await message.answer(f"⚠️ Задача создана, но не удалось отправить исполнителю")
    
    except Exception as e:
        logger.error(f"Ошибка создания задачи: {e}")
        await message.answer(f"❌ Ошибка создания задачи: {e}")

# ==================== ОБРАБОТЧИКИ CALLBACK ====================

@dp.callback_query(F.data.startswith('status_'))
async def process_status_change(callback: types.CallbackQuery):
    """Обработка изменения статуса задачи"""
    parts = callback.data.replace('status_', '').split('_', 1)
    if len(parts) < 2:
        await callback.answer("❌ Ошибка", show_alert=True)
        return
    
    new_status, task_number = parts
    
    task = db.get_task_by_number(task_number)
    if not task:
        await callback.answer("❌ Задача не найдена", show_alert=True)
        return
    
    # Проверяем права
    if task['tm_id'] != callback.from_user.id:
        await callback.answer("❌ Это не ваша задача", show_alert=True)
        return
    
    # Проверяем подзадачи перед завершением
    if new_status == 'Выполнено':
        subtasks = db.get_subtasks(task_number)
        if subtasks and not all(s['is_done'] for s in subtasks):
            await callback.answer("❌ Сначала завершите все подзадачи", show_alert=True)
            return
    
    # Обновляем статус
    if db.update_task_status(task_number, new_status, callback.from_user.id):
        # Уведомляем админа
        user = db.get_user_by_tg_id(callback.from_user.id)
        username = user['username'] if user else 'Неизвестный'
        
        await notify_admin(
            f"📊 <b>Статус задачи #{task_number} изменен</b>\n\n"
            f"Исполнитель: {username}\n"
            f"Новый статус: {new_status}\n\n"
            f"Задача: {task['task_text'][:100]}..."
        )
        
        # Обновляем сообщение
        updated_task = db.get_task_by_number(task_number)
        msg_text, keyboard = format_task_message(updated_task)
        
        try:
            await callback.message.edit_text(msg_text, reply_markup=keyboard, parse_mode='HTML')
        except:
            await callback.message.answer(msg_text, reply_markup=keyboard, parse_mode='HTML')
        
        await callback.answer(f"✅ Статус изменен на: {new_status}")
    else:
        await callback.answer("❌ Ошибка изменения статуса", show_alert=True)

@dp.callback_query(F.data.startswith('comment_'))
async def process_comment_request(callback: types.CallbackQuery, state: FSMContext):
    """Запрос комментария"""
    task_number = callback.data.replace('comment_', '')
    
    task = db.get_task_by_number(task_number)
    if not task:
        await callback.answer("❌ Задача не найдена", show_alert=True)
        return
    
    await state.update_data(task_number=task_number, admin_comment=False)
    await state.set_state(TaskStates.waiting_for_comment)
    
    await callback.message.answer(
        f"💬 <b>Комментарий к задаче #{task_number}</b>\n\n"
        f"Напишите ваш комментарий:\n"
        f"(будет отправлен администратору)\n\n"
        f"Или /cancel для отмены",
        parse_mode='HTML'
    )
    await callback.answer()

@dp.callback_query(F.data.startswith('subtasks_'))
async def process_subtasks_view(callback: types.CallbackQuery):
    """Просмотр и управление подзадачами"""
    task_number = callback.data.replace('subtasks_', '')
    
    subtasks = db.get_subtasks(task_number)
    if not subtasks:
        await callback.answer("❌ Подзадачи не найдены", show_alert=True)
        return
    
    # Формируем список подзадач с кнопками
    buttons = []
    for subtask in subtasks:
        checkbox = "✅" if subtask['is_done'] else "☐"
        buttons.append([
            InlineKeyboardButton(
                text=f"{checkbox} {subtask['text'][:40]}",
                callback_data=f"toggle_subtask_{subtask['id']}"
            )
        ])
    
    buttons.append([InlineKeyboardButton(text="🔙 Назад к задаче", callback_data=f"back_to_task_{task_number}")])
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
    
    await callback.message.edit_text(
        f"📋 <b>Подзадачи #{task_number}</b>\n\n"
        f"Нажмите на подзадачу чтобы отметить выполненной:",
        reply_markup=keyboard,
        parse_mode='HTML'
    )
    await callback.answer()

@dp.callback_query(F.data.startswith('toggle_subtask_'))
async def process_toggle_subtask(callback: types.CallbackQuery):
    """Переключение статуса подзадачи"""
    subtask_id = int(callback.data.replace('toggle_subtask_', ''))
    
    if db.toggle_subtask(subtask_id):
        # Получаем обновленный список подзадач
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT task_number FROM subtasks WHERE id = ?', (subtask_id,))
        row = cursor.fetchone()
        conn.close()
        
        if row:
            task_number = row['task_number']
            subtasks = db.get_subtasks(task_number)
            
            # Обновляем кнопки
            buttons = []
            for subtask in subtasks:
                checkbox = "✅" if subtask['is_done'] else "☐"
                buttons.append([
                    InlineKeyboardButton(
                        text=f"{checkbox} {subtask['text'][:40]}",
                        callback_data=f"toggle_subtask_{subtask['id']}"
                    )
                ])
            
            buttons.append([InlineKeyboardButton(text="🔙 Назад к задаче", callback_data=f"back_to_task_{task_number}")])
            
            keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
            
            await callback.message.edit_reply_markup(reply_markup=keyboard)
            await callback.answer("✅ Статус подзадачи изменен")
        else:
            await callback.answer("❌ Ошибка", show_alert=True)
    else:
        await callback.answer("❌ Ошибка изменения статуса", show_alert=True)

@dp.callback_query(F.data.startswith('back_to_task_'))
async def process_back_to_task(callback: types.CallbackQuery):
    """Возврат к просмотру задачи"""
    task_number = callback.data.replace('back_to_task_', '')
    
    task = db.get_task_by_number(task_number)
    if task:
        msg_text, keyboard = format_task_message(task)
        await callback.message.edit_text(msg_text, reply_markup=keyboard, parse_mode='HTML')
    else:
        await callback.answer("❌ Задача не найдена", show_alert=True)

@dp.callback_query(F.data == "admin_cancel")
async def process_admin_cancel(callback: types.CallbackQuery, state: FSMContext):
    """Отмена действия админа"""
    await state.clear()
    await callback.message.edit_text("❌ Действие отменено")
    await callback.answer()

@dp.callback_query(F.data.startswith('admin_'))
async def process_admin_action(callback: types.CallbackQuery, state: FSMContext):
    """Обработка действий администратора"""
    action_data = callback.data.split('_')
    action = action_data[1]
    
    if action == 'cancel':
        await callback.answer()
        return
    
    task_number = '_'.join(action_data[2:]) if len(action_data) > 2 else None
    
    if not task_number:
        data = await state.get_data()
        task_number = data.get('selected_task')
    
    if not task_number:
        await callback.answer("❌ Задача не найдена", show_alert=True)
        return
    
    task = db.get_task_by_number(task_number)
    if not task:
        await callback.answer("❌ Задача не найдена", show_alert=True)
        await state.clear()
        return
    
    if action == 'edit':
        # Редактирование задачи
        await state.update_data(task_number=task_number)
        await state.set_state(TaskStates.waiting_for_edit)
        
        await callback.message.answer(
            f"📝 <b>Редактирование задачи #{task_number}</b>\n\n"
            f"Текущие данные:\n"
            f"Приоритет: {task['priority']}\n"
            f"Задача: {task['task_text']}\n"
            f"Дедлайн: {task['deadline']}\n"
            f"Комментарий: {task.get('comment', '-')}\n\n"
            f"Отправьте новое описание в формате:\n\n"
            f"Приоритет: Высокий/Средний/Низкий\n"
            f"Задача: Текст задачи\n"
            f"Дедлайн: ДД.ММ.ГГГГ ЧЧ:ММ или -\n"
            f"Комментарий: Текст\n\n"
            f"Или /cancel для отмены",
            parse_mode='HTML'
        )
        await callback.answer()
    
    elif action == 'delete':
        # Удаление задачи
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="✅ Да, удалить", callback_data=f"confirm_delete_{task_number}"),
                InlineKeyboardButton(text="❌ Отмена", callback_data="admin_cancel")
            ]
        ])
        
        await callback.message.edit_text(
            f"⚠️ <b>Удаление задачи #{task_number}</b>\n\n"
            f"Вы уверены что хотите удалить эту задачу?\n"
            f"Это действие нельзя отменить!\n\n"
            f"Задача: {task['task_text'][:100]}",
            reply_markup=keyboard,
            parse_mode='HTML'
        )
        await callback.answer()
    
    elif action == 'comment':
        # Комментарий к задаче
        await state.update_data(task_number=task_number, admin_comment=True)
        await state.set_state(TaskStates.waiting_for_comment)
        
        await callback.message.answer(
            f"💬 <b>Комментарий к задаче #{task_number}</b>\n\n"
            f"Напишите ваш комментарий:\n"
            f"(будет отправлен исполнителю)\n\n"
            f"Или /cancel для отмены",
            parse_mode='HTML'
        )
        await callback.answer()
    
    elif action == 'history':
        # История задачи
        history = db.get_task_history(task_number)
        
        if not history:
            await callback.answer("История пуста", show_alert=True)
            return
        
        response = f"📜 <b>История задачи #{task_number}</b>\n\n"
        for entry in history:
            timestamp = entry['timestamp'][:16] if len(entry['timestamp']) > 16 else entry['timestamp']
            username = entry.get('changed_by_username', 'Система')
            response += f"🕒 {timestamp} | {username}\n"
            response += f"▪️ {entry['action']}\n"
            if entry.get('comment'):
                response += f"💬 {entry['comment']}\n"
            response += "\n"
        
        # Кнопка назад
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔙 Назад", callback_data="admin_cancel")]
        ])
        
        await callback.message.edit_text(response, reply_markup=keyboard, parse_mode='HTML')
        await callback.answer()
    
    elif action == 'status':
        # Изменение статуса
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="🆕 Новая", callback_data=f"set_status_Новая_{task_number}"),
                InlineKeyboardButton(text="⏳ В работе", callback_data=f"set_status_В работе_{task_number}")
            ],
            [
                InlineKeyboardButton(text="✅ Выполнено", callback_data=f"set_status_Выполнено_{task_number}"),
                InlineKeyboardButton(text="❌ Не выполнено", callback_data=f"set_status_Не выполнено_{task_number}")
            ],
            [
                InlineKeyboardButton(text="⚠️ Просрочена", callback_data=f"set_status_Просрочена_{task_number}"),
                InlineKeyboardButton(text="🔙 Отмена", callback_data="admin_cancel")
            ]
        ])
        
        await callback.message.edit_text(
            f"🔄 <b>Изменение статуса задачи #{task_number}</b>\n\n"
            f"Текущий статус: {task['status']}\n\n"
            f"Выберите новый статус:",
            reply_markup=keyboard,
            parse_mode='HTML'
        )
        await callback.answer()

@dp.callback_query(F.data.startswith('confirm_delete_'))
async def confirm_delete_task(callback: types.CallbackQuery, state: FSMContext):
    """Подтверждение удаления задачи"""
    task_number = callback.data.replace('confirm_delete_', '')
    
    task = db.get_task_by_number(task_number)
    if not task:
        await callback.answer("❌ Задача не найдена", show_alert=True)
        return
    
    # Удаляем задачу
    conn = db.get_connection()
    cursor = conn.cursor()
    
    try:
        # Удаляем подзадачи
        cursor.execute('DELETE FROM subtasks WHERE task_number = ?', (task_number,))
        
        # Удаляем историю
        cursor.execute('DELETE FROM task_history WHERE task_number = ?', (task_number,))
        
        # Удаляем саму задачу
        cursor.execute('DELETE FROM tasks WHERE task_number = ?', (task_number,))
        
        conn.commit()
        
        # Уведомляем исполнителя
        try:
            await bot.send_message(
                task['tm_id'],
                f"🗑 <b>Задача #{task_number} удалена администратором</b>\n\n"
                f"Задача: {task['task_text'][:100]}",
                parse_mode='HTML'
            )
        except Exception as e:
            logger.error(f"Ошибка уведомления при удалении: {e}")
        
        await callback.message.edit_text(
            f"✅ Задача #{task_number} успешно удалена"
        )
        await state.clear()
        await callback.answer()
        
    except Exception as e:
        logger.error(f"Ошибка удаления задачи: {e}")
        await callback.answer("❌ Ошибка при удалении задачи", show_alert=True)
        conn.rollback()
    finally:
        conn.close()

@dp.callback_query(F.data.startswith('set_status_'))
async def set_task_status(callback: types.CallbackQuery, state: FSMContext):
    """Установка статуса задачи администратором"""
    parts = callback.data.replace('set_status_', '').split('_', 1)
    if len(parts) < 2:
        await callback.answer("❌ Ошибка", show_alert=True)
        return
    
    new_status, task_number = parts
    
    task = db.get_task_by_number(task_number)
    if not task:
        await callback.answer("❌ Задача не найдена", show_alert=True)
        return
    
    # Обновляем статус
    db.update_task_status(task_number, new_status, callback.from_user.id)
    
    # Уведомляем исполнителя
    try:
        await bot.send_message(
            task['tm_id'],
            f"🔄 <b>Статус задачи #{task_number} изменен</b>\n\n"
            f"Новый статус: {new_status}\n"
            f"Изменил: администратор",
            parse_mode='HTML'
        )
    except Exception as e:
        logger.error(f"Ошибка уведомления при изменении статуса: {e}")
    
    await callback.message.edit_text(
        f"✅ Статус задачи #{task_number} изменен на: {new_status}"
    )
    await state.clear()
    await callback.answer()

@dp.callback_query(F.data == "noop")
async def process_noop(callback: types.CallbackQuery):
    """Обработка пустого callback"""
    await callback.answer()

# ==================== ЗАПУСК БОТА ====================

async def main():
    """Главная функция запуска бота"""
    global ADMIN_ID
    
    logger.info("Бот запускается...")
    
    # Пытаемся найти админа в базе
    ADMIN_ID = db.get_admin_id()
    
    if ADMIN_ID:
        logger.info(f"Admin ID найден в базе: {ADMIN_ID}")
    else:
        logger.info("Ожидается первое сообщение от администратора")
    
    # Запуск фоновых задач
    asyncio.create_task(check_deadlines())
    asyncio.create_task(send_daily_report())
    
    logger.info("Бот готов к работе!")
    
    # Запуск бота
    try:
        await dp.start_polling(bot, skip_updates=True)
    except Exception as e:
        logger.error(f"Ошибка при запуске бота: {e}")
        raise

if __name__ == '__main__':
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("Бот остановлен пользователем")
    except Exception as e:
        logger.error(f"Критическая ошибка: {e}")
